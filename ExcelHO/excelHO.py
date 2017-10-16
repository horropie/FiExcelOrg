'''
Script by Jakob Hollweck
This only categorizes the cells and sums them up.
'''



import openpyxl
from operator import is_not
wb=openpyxl.load_workbook("Finanzen 2017.xlsx")
mysheet=wb.get_sheet_by_name("Juni")
#for col in mysheet.iter_cols(min_row=15, min_col=4, max_col=1, max_row=100):
#    for cell in col:
#        if mysheet[str(cell)].value==

Unterkunft_trigger=["unt."]
Lebensmittel_trigger=["lem.","Aldi", "aldi", "Lidl", "lidl", "Bioladen", "bioladen"]
Restaurant_trigger=["rtr.","mensa", "Mensa"]
Aktivitäten_trigger=["akt.", "Bier", "Alkohol"]
Kleidung_trigger=["kle."]
Gesundheit_trigger=["ges."]
Transport_trigger=["tra."]
Sonstiges_trigger=["son."]


#the following is quite ugly rn and definitely should be shorter

s="=SUM("
for j in range(15, 150):
    e=mysheet.cell(row=j, column=4)
    mysheet["C7"]=s
    if e.value in str(Unterkunft_trigger):
    #if e.value is "u":
        s+="C"+str(j)+","
s2=s
s2+=")"
mysheet["C7"]=s2

s="=SUM("
for j in range(15, 100):
    e=mysheet.cell(row=j, column=4)
    mysheet["C8"]=s
    if e.value in str(Lebensmittel_trigger):
    #if e.value is "l":
        s+="C"+str(j)+","

s2=s
s2+=")"
mysheet["C8"]=s2

s="=SUM("
for j in range(15, 100):
    e=mysheet.cell(row=j, column=4)
    mysheet["C9"]=s
    if e.value in str(Restaurant_trigger):
    #if e.value is "r":
        s+="C"+str(j)+","

s2=s
s2+=")"
mysheet["C9"]=s2

s="=SUM("
for j in range(15, 100):
    e=mysheet.cell(row=j, column=4)
    mysheet["C10"]=s
    if e.value in str(Aktivitäten_trigger):
    #if e.value is "a":
        s+="C"+str(j)+","

s2=s
s2+=")"
mysheet["C10"]=s2


s="=SUM("
for j in range(15, 100):
    e=mysheet.cell(row=j, column=4)
    mysheet["C11"]=s
    if e.value in str(Kleidung_trigger):
    #if e.value is "k":
        s+="C"+str(j)+","

s2=s
s2+=")"
mysheet["C11"]=s2

s="=SUM("
for j in range(15, 100):
    e=mysheet.cell(row=j, column=4)
    mysheet["C12"]=s
    if e.value in str(Gesundheit_trigger):
    #if e.value is "h":
        s+="C"+str(j)+","

s2=s
s2+=")"
mysheet["C12"]=s2

s="=SUM("
for j in range(15, 100):
    e=mysheet.cell(row=j, column=4)
    mysheet["C13"]=s
    if e.value in str(Transport_trigger):
    #if e.value is "t":
        s+="C"+str(j)+","

s2=s
s2+=")"
mysheet["C13"]=s2

s="=SUM("
for j in range(15, 100):
    e=mysheet.cell(row=j, column=4)
    mysheet["C14"]=s
    if e.value in str(Sonstiges_trigger):
    #if e.value is "s":
        s+="C"+str(j)+","

s2=s
s2+=")"
mysheet["C14"]=s2
#would be nicer to have several loops in a loop with "else" for "Sonstiges"
#furthermore a custom-function containing all the bs would be nice

print(mysheet["C7"].value)
print(mysheet["C8"].value)
print(mysheet["C9"].value)
print(mysheet["C10"].value)
print(mysheet["C11"].value)
print(mysheet["C12"].value)
print(mysheet["C13"].value)
print(mysheet["C14"].value)

print(mysheet["B22"].value)
print(str(Lebensmittel_trigger))


wb.save('Finanzen 2017_Test.xlsx')
