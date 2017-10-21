'''
Script by Jakob Hollweck
This only categorizes the cells and sums them up.
'''

#-------------------------------------------------------------------------------

import openpyxl
from operator import is_not
wb=openpyxl.load_workbook("Finanzen 2017.xlsx")
mysheet=wb.get_sheet_by_name("Juni")

#-------------------------------------------------------------------------------

Unterkunft_trigger=["unt.", "Miete", "miete"]
Lebensmittel_trigger=["lem.","Aldi", "aldi", "Lidl", "lidl", "tegut", "Tegut", "Bäcker", "Backwerk", "Bioladen", "bioladen",
"Brezen", "Semmel", "Brot",]
Restaurant_trigger=["rtr.","mensa", "Mensa", "Essen", "essen", "Eis", "eis"]
Aktivitäten_trigger=["akt.", "Bier", "Alkohol"]
Kleidung_trigger=["kle."]
Gesundheit_trigger=["ges."]
Transport_trigger=["tra.", "Parken", "parken", "Parkplatz", "parkplatz"]
Sonstiges_trigger=["son.", "Kopierer", "kopierer", "Block", "Waschen", "Drucken", "Easybell"]


#the following is quite ugly rn and definitely should be shorter
#-------------------------------------------------------------------------------

def categorizer_col(cell, *list, rangestart=15, rangeend=150, column=2, inp="C"):
    s="=SUM("
    for j in range(rangestart, rangeend):
        e=mysheet.cell(row=j, column)
        if type(e.value) is str:
            mysheet[cell]=s
            if e.value in str(*list):
                s+=inp + str(j)+","
    s2=s
    s2+=")"
    mysheet[cell]=s2


categorizer_col("C7", Unterkunft_trigger)
print(mysheet["C7"].value)

#-------------------------------------------------------------------------------

s="=SUM("
for j in range(15, 100):
    e=mysheet.cell(row=j, column=2)
    if type(e.value) is str:
        mysheet["C8"]=s
        if e.value in str(Lebensmittel_trigger):
            s+="C"+str(j)+","
s2=s
s2+=")"
mysheet["C8"]=s2

#-------------------------------------------------------------------------------

s="=SUM("
for j in range(15, 100):
    e=mysheet.cell(row=j, column=2)
    if type(e.value) is str:
        mysheet["C9"]=s
        if e.value in str(Restaurant_trigger):
            s+="C"+str(j)+","
s2=s
s2+=")"
mysheet["C9"]=s2

#-------------------------------------------------------------------------------

s="=SUM("
for j in range(15, 100):
    e=mysheet.cell(row=j, column=2)
    if type(e.value) is str:
        mysheet["C10"]=s
        if e.value in str(Aktivitäten_trigger):
            s+="C"+str(j)+","
s2=s
s2+=")"
mysheet["C10"]=s2

#-------------------------------------------------------------------------------

s="=SUM("
for j in range(15, 100):
    e=mysheet.cell(row=j, column=2)
    if type(e.value) is str:
        mysheet["C11"]=s
        if e.value in str(Kleidung_trigger):
            s+="C"+str(j)+","
s2=s
s2+=")"
mysheet["C11"]=s2

#-------------------------------------------------------------------------------

s="=SUM("
for j in range(15, 100):
    e=mysheet.cell(row=j, column=2)
    if type(e.value) is str:
        mysheet["C12"]=s
        if e.value in str(Gesundheit_trigger):
            s+="C"+str(j)+","
s2=s
s2+=")"
mysheet["C12"]=s2

#-------------------------------------------------------------------------------

s="=SUM("
for j in range(15, 100):
    e=mysheet.cell(row=j, column=2)
    if type(e.value) is str:
        mysheet["C13"]=s
        if e.value in str(Transport_trigger):
            s+="C"+str(j)+","
s2=s
s2+=")"
mysheet["C13"]=s2

#-------------------------------------------------------------------------------

s="=SUM("
for j in range(15, 100):
    e=mysheet.cell(row=j, column=2)
    if type(e.value) is str:
        mysheet["C14"]=s
        if e.value in str(Sonstiges_trigger):
            s+="C"+str(j)+","
s2=s
s2+=")"
mysheet["C14"]=s2

#-------------------------------------------------------------------------------

#would be nicer to have several loops in a loop with "else" for "Sonstiges"
#furthermore a custom-function containing all the bs would be nice

#-------------------------------------------------------------------------------
'''
print(mysheet["C7"].value)
print(mysheet["C8"].value)
print(mysheet["C9"].value)
print(mysheet["C10"].value)
print(mysheet["C11"].value)
print(mysheet["C12"].value)
print(mysheet["C13"].value)
print(mysheet["C14"].value)
'''
#-------------------------------------------------------------------------------
wb.save('Finanzen 2017_Test.xlsx')
