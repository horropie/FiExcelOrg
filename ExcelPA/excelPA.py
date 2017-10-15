import openpyxl
excel_document = openpyxl.load_workbook('sample.xlsx')

#typeofthesheet
print(type(excel_document))

#openingthesheet
print(excel_document.get_sheet_names())
sheet1=excel_document.get_sheet_by_name('Sheet1')

#accessing values in certain cells
print(sheet1['A2'].value)
print(sheet1.cell(row=5, column=2).value)

#Celltype
print(type(sheet1['A2']))
#or
print(sheet1.cell(row=5,column=2))

#multiplecells
multiple_cells=sheet1["A1":"B3"]
for row in multiple_cells:
    for cell in row:
        print(cell.value)

#rows and columns accessing
all_rows = sheet1.rows
print(all_rows[:])
all_columns = sheet1.columns
print(all_columns[:])
