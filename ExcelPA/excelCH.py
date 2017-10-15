import openpyxl
wb=openpyxl.load_workbook("cr1.xlsx")
sheet=wb.active
sheet.title="Working on Save as"
wb.create_sheet(index=0, title="Slide1")
wb.create_sheet(index=2, title="Slide3")
wb.get_sheet_names()
wb.remove_sheet(wb.get_sheet_by_name("Slide3"))
wb.get_sheet_names()
mysheet=wb.get_sheet_by_name("Slide1")
mysheet["F6"]="Writing new Value"
mysheet["F6"].value

mysheet.merge_cells('B2:D3')
mysheet['A1']='cells merged together'
mysheet.merge_cells('F6:F7')
mysheet['G5']='Two merged cells.'
wb.save('Mergingcells.xlsx')

wb.save('cr1_2.xlsx')
