import openpyxl
wb=openpyxl.load_workbook("cr1.xlsx")
mysheet=wb.active


mysheet.merge_cells('B2:D3')
mysheet['A1']='cells merged together'
mysheet.merge_cells('F6:F7')
mysheet['G5']='Two merged cells.'
wb.save('Mergingcells.xlsx')
